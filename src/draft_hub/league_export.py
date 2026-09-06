"""League workbook (Excel) — current rosters, salaries, and history."""

from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.draft_hub import storage

_FILE_PART_RE = re.compile(r"[^\w\-]+", re.UNICODE)


def league_name_matches(typed: str, official: str) -> bool:
    left = " ".join(str(typed or "").split()).casefold()
    right = " ".join(str(official or "").split()).casefold()
    return bool(left) and left == right


def workbook_filename(league: dict[str, Any]) -> str:
    slug = _FILE_PART_RE.sub("-", str(league.get("name") or "league").strip()).strip("-")
    slug = (slug or "league")[:48]
    season = league.get("season") or ""
    return f"ScoreSense-{slug}-{season}-league.xlsx"


def build_league_workbook(league_id: str) -> tuple[bytes, str]:
    league = storage.get_league(league_id)
    if not league:
        raise ValueError("League not found")
    teams = storage.list_league_teams(league_id)
    by_team = storage.list_league_rosters_by_team(league_id)
    contracts = storage.list_league_contract_rows(league_id)
    trades = _list_trade_log(league_id)
    proposals = storage.list_trade_proposals(league_id)
    team_scores = _list_rows(
        "SELECT season, week, team_id, matchup_id, points, scored_at "
        "FROM league_team_week_score WHERE league_id = ? ORDER BY season, week, team_id",
        (league_id,),
    )
    player_scores = _list_rows(
        "SELECT season, week, team_id, player_id, slot, "
        "lineup_role, points, scored_at "
        "FROM league_player_week_score WHERE league_id = ? "
        "ORDER BY season, week, team_id, player_id",
        (league_id,),
    )
    sleeper_weeks = _sleeper_week_rows(league)

    team_label = {
        str(t["id"]): (t.get("owner_name") or t.get("name") or t["id"]) for t in teams
    }
    seat_label = {str(t["id"]): t.get("name") or "" for t in teams}

    wb = Workbook()
    _write_league_sheet(wb.active, league, teams)
    _write_sheet(
        wb.create_sheet("Teams"),
        ["Owner", "Seat", "Commissioner", "Cap remaining", "Sleeper roster"],
        [
            [
                t.get("owner_name") or t.get("name") or "",
                t.get("name") or "",
                "yes" if t.get("is_commissioner") else "",
                t.get("budget_remaining"),
                t.get("sleeper_team_name") or t.get("sleeper_roster_id") or "",
            ]
            for t in teams
        ],
    )
    roster_rows: list[list[Any]] = []
    for team in teams:
        tid = str(team["id"])
        for row in by_team.get(tid, []):
            roster_rows.append(
                [
                    team_label.get(tid, ""),
                    seat_label.get(tid, ""),
                    row.get("player_name") or "",
                    row.get("team") or "",
                    row.get("position") or "",
                    row.get("salary"),
                    row.get("contract_years"),
                    row.get("roster_status") or "",
                    (row.get("contract") or {}).get("contract_type") or "",
                    row.get("acquired_at") or "",
                ]
            )
    _write_sheet(
        wb.create_sheet("Rosters"),
        [
            "Owner",
            "Seat",
            "Player",
            "NFL",
            "Pos",
            "Salary",
            "Years",
            "Status",
            "Type",
            "Acquired",
        ],
        roster_rows,
    )
    _write_sheet(
        wb.create_sheet("Salary history"),
        [
            "Season",
            "Owner",
            "Seat",
            "Player",
            "Pos",
            "Base salary",
            "Cap hit",
            "Status",
            "Phase",
            "Acquisition",
            "Source",
        ],
        [
            [
                row.get("season_year"),
                row.get("owner_label") or "",
                row.get("hub_team_name") or "",
                row.get("player_name") or "",
                row.get("position") or "",
                row.get("base_salary"),
                row.get("cap_hit"),
                row.get("roster_status") or "",
                row.get("contract_phase") or "",
                row.get("acquisition_type") or "",
                row.get("source_kind") or "",
            ]
            for row in contracts
        ],
    )
    _write_sheet(
        wb.create_sheet("Trades"),
        ["When", "Side A", "Side B", "Send A", "Send B"],
        [
            [
                row.get("created_at") or "",
                team_label.get(str(row.get("team_a_id") or ""), row.get("team_a_id") or ""),
                team_label.get(str(row.get("team_b_id") or ""), row.get("team_b_id") or ""),
                _json_cell(row.get("send_a")),
                _json_cell(row.get("send_b")),
            ]
            for row in trades
        ],
    )
    _write_sheet(
        wb.create_sheet("Trade proposals"),
        ["When", "Status", "Note", "Parties"],
        [
            [
                row.get("created_at") or "",
                row.get("status") or "",
                row.get("note") or "",
                _json_cell(row.get("parties")),
            ]
            for row in proposals
        ],
    )
    _write_sheet(
        wb.create_sheet("Week scores"),
        ["Season", "Week", "Owner", "Matchup", "Points", "Scored at"],
        [
            [
                row.get("season"),
                row.get("week"),
                team_label.get(str(row.get("team_id") or ""), row.get("team_id") or ""),
                row.get("matchup_id") or "",
                row.get("points"),
                row.get("scored_at") or "",
            ]
            for row in team_scores
        ],
    )
    _write_sheet(
        wb.create_sheet("Player week scores"),
        ["Season", "Week", "Owner", "Player", "Slot", "Role", "Points"],
        [
            [
                row.get("season"),
                row.get("week"),
                team_label.get(str(row.get("team_id") or ""), row.get("team_id") or ""),
                row.get("player_id") or "",
                row.get("slot") or "",
                row.get("lineup_role") or "",
                row.get("points"),
            ]
            for row in player_scores
        ],
    )
    _write_sheet(
        wb.create_sheet("Scoring history"),
        ["Season", "Week", "Roster", "Points", "Playoff"],
        sleeper_weeks,
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), workbook_filename(league)


def _write_league_sheet(ws, league: dict[str, Any], teams: list[dict[str, Any]]) -> None:
    ws.title = "League"
    rules = league.get("rules") or {}
    rows = [
        ["Name", league.get("name") or ""],
        ["Season", league.get("season")],
        ["Room", league.get("room_code") or ""],
        ["Status", league.get("status") or ""],
        ["Seats", league.get("team_count")],
        ["Managers", len(teams)],
        ["Salary cap", (rules or {}).get("salary_cap")],
        ["Scoring", (rules or {}).get("scoring") or ""],
        ["Draft completed", "yes" if league.get("draft_completed") else ""],
        ["Exported for Excel", "Rosters, salaries, trades, and week history"],
    ]
    _write_sheet(ws, ["Field", "Value"], rows)


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    bold = Font(bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = bold
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, _cell_value(value))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
    for col, header in enumerate(headers, start=1):
        width = len(str(header))
        for row in rows[:80]:
            width = max(width, len(str(_cell_value(row[col - 1]) if col - 1 < len(row) else "")))
        ws.column_dimensions[get_column_letter(col)].width = min(40, max(12, width + 2))


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else ""
    if isinstance(value, (dict, list)):
        return _json_cell(value)
    return value


def _json_cell(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=True, default=str)
    except TypeError:
        return str(value)


def _list_trade_log(league_id: str) -> list[dict[str, Any]]:
    with storage.get_conn() as conn:
        rows = conn.execute(
            """SELECT league_id, team_a_id, team_b_id, send_a_json, send_b_json, created_at
               FROM trade_log WHERE league_id = ? ORDER BY created_at""",
            (league_id,),
        ).fetchall()
    out = []
    for row in rows:
        send_a = row["send_a_json"]
        send_b = row["send_b_json"]
        try:
            send_a = json.loads(send_a) if send_a else ""
        except json.JSONDecodeError:
            pass
        try:
            send_b = json.loads(send_b) if send_b else ""
        except json.JSONDecodeError:
            pass
        out.append(
            {
                "team_a_id": row["team_a_id"],
                "team_b_id": row["team_b_id"],
                "send_a": send_a,
                "send_b": send_b,
                "created_at": row["created_at"],
            }
        )
    return out


def _list_rows(sql: str, params: tuple[Any, ...]) -> list[dict[str, Any]]:
    with storage.get_conn() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def _sleeper_week_rows(league: dict[str, Any]) -> list[list[Any]]:
    sleeper_id = str(league.get("sleeper_league_id") or "").strip()
    if not sleeper_id:
        return []
    cached = storage.get_sleeper_scoring_cache(sleeper_id)
    if not cached:
        return []
    payload = cached.get("payload") or {}
    season = payload.get("season") or league.get("season")
    out: list[list[Any]] = []
    for week in payload.get("weeks") or []:
        week_no = week.get("week")
        playoff = "yes" if week.get("is_playoff") else ""
        matchups = week.get("matchups") or week.get("rosters") or []
        if isinstance(matchups, dict):
            matchups = list(matchups.values())
        for item in matchups:
            if not isinstance(item, dict):
                continue
            out.append(
                [
                    season,
                    week_no,
                    item.get("owner")
                    or item.get("display_name")
                    or item.get("roster_id")
                    or "",
                    item.get("points") if item.get("points") is not None else item.get("fpts"),
                    playoff,
                ]
            )
    return out
