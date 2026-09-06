"""League workbook export and multi-commissioner delete."""

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api import app
from app.auth import require_hub_user
from src.draft_hub import storage
from src.draft_hub.league_export import build_league_workbook, league_name_matches
from src.draft_hub.league_home import build_league_home
from src.draft_hub.league_lifecycle import (
    approve_league_delete,
    cancel_league_delete,
    delete_request_snapshot,
    start_league_delete,
)
from src.draft_hub.schemas import LeagueRules


def _client_for(sub: str) -> TestClient:
    app.dependency_overrides[require_hub_user] = lambda: {"sub": sub, "auth_type": "dev"}
    return TestClient(app)


def teardown_function() -> None:
    app.dependency_overrides.pop(require_hub_user, None)


def _seed_league(name="Lifecycle League", comm="life-comm"):
    rules = LeagueRules()
    league = storage.create_league(comm, name, 2026, rules, team_count=10)
    comm_team = storage.get_team_by_user(league["id"], comm)
    return league, comm_team


def test_league_name_matches_ignores_case_and_space():
    assert league_name_matches("  My Auction  ", "My Auction")
    assert not league_name_matches("My", "My Auction")
    assert not league_name_matches("", "My Auction")


def test_member_can_export_workbook(hub_db):
    league, comm_team = _seed_league()
    ws = storage.roster_workspace_for_league(league)
    storage.add_roster_slot(
        ws,
        {
            "player_id": "00-0033873",
            "player_name": "Patrick Mahomes",
            "team": "KC",
            "position": "QB",
            "salary": 45,
            "contract_years": 3,
        },
        team_id=comm_team["id"],
    )
    storage.replace_league_contract_season_source(
        league["id"],
        2025,
        [
            {
                "owner_label": "Comm",
                "hub_team_name": comm_team["name"],
                "player_name": "Patrick Mahomes",
                "player_id": "00-0033873",
                "position": "QB",
                "base_salary": 40,
                "cap_hit": 40,
                "roster_status": "active",
            }
        ],
        source_kind="import",
    )
    payload, filename = build_league_workbook(league["id"])
    assert filename.endswith(".xlsx")
    assert "Lifecycle" in filename
    book = load_workbook(BytesIO(payload))
    assert "Rosters" in book.sheetnames
    assert "Salary history" in book.sheetnames
    roster_values = [cell.value for row in book["Rosters"].iter_rows(min_row=2) for cell in row]
    assert "Patrick Mahomes" in roster_values
    history_values = [cell.value for row in book["Salary history"].iter_rows(min_row=2) for cell in row]
    assert "Patrick Mahomes" in history_values

    member = storage.join_league("life-member", league["room_code"], "Member Seat")
    client = _client_for("life-member")
    res = client.get(f"/api/hub/league/{league['id']}/export")
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    outsider = _client_for("life-stranger")
    denied = outsider.get(f"/api/hub/league/{league['id']}/export")
    assert denied.status_code == 403
    assert member["id"]


def test_single_commissioner_delete_needs_name(hub_db):
    league, _team = _seed_league("Solo Delete")
    snap = delete_request_snapshot(league["id"], viewer_sub="life-comm")
    assert snap["required_count"] == 1
    assert snap["can_start"] is True
    try:
        start_league_delete(league["id"], actor_sub="life-comm", confirm_name="Wrong")
        assert False, "expected name mismatch"
    except Exception as exc:
        assert "league name" in str(exc).lower()
    result = start_league_delete(league["id"], actor_sub="life-comm", confirm_name="Solo Delete")
    assert result["deleted"] is True
    assert storage.get_league(league["id"]) is None


def test_all_commissioners_must_approve(hub_db):
    league, _team = _seed_league("Shared Delete")
    member = storage.join_league("life-co", league["room_code"], "Co Seat")
    storage.set_team_co_commissioner(
        league["id"], member["id"], enabled=True, actor_sub="life-comm"
    )
    started = start_league_delete(league["id"], actor_sub="life-comm", confirm_name="Shared Delete")
    assert started["deleted"] is False
    assert started["approved_count"] == 1
    assert started["required_count"] == 2
    assert storage.get_league(league["id"]) is not None

    client = _client_for("life-co")
    status = client.get(f"/api/hub/league/{league['id']}/delete-request")
    assert status.status_code == 200
    body = status.json()
    assert body["pending"] is True
    assert body["you_approved"] is False

    member_client = _client_for("life-fan")
    storage.join_league("life-fan", league["room_code"], "Fan Seat")
    forbidden = member_client.post(
        f"/api/hub/league/{league['id']}/delete-request/approve",
        json={"confirm_name": "Shared Delete"},
    )
    assert forbidden.status_code == 403

    finished = approve_league_delete(
        league["id"], actor_sub="life-co", confirm_name="Shared Delete"
    )
    assert finished["deleted"] is True
    assert storage.get_league(league["id"]) is None


def test_new_commissioner_must_also_agree(hub_db):
    league, _team = _seed_league("Growing Delete")
    first = storage.join_league("life-co-a", league["room_code"], "Co A")
    storage.set_team_co_commissioner(
        league["id"], first["id"], enabled=True, actor_sub="life-comm"
    )
    start_league_delete(league["id"], actor_sub="life-comm", confirm_name="Growing Delete")
    second = storage.join_league("life-co-b", league["room_code"], "Co B")
    storage.set_team_co_commissioner(
        league["id"], second["id"], enabled=True, actor_sub="life-comm"
    )
    mid = approve_league_delete(
        league["id"], actor_sub="life-co-a", confirm_name="Growing Delete"
    )
    assert mid["deleted"] is False
    assert mid["required_count"] == 3
    last = approve_league_delete(
        league["id"], actor_sub="life-co-b", confirm_name="Growing Delete"
    )
    assert last["deleted"] is True


def test_cancel_withdraws_delete(hub_db):
    league, _team = _seed_league("Withdraw Delete")
    member = storage.join_league("life-co-c", league["room_code"], "Co C")
    storage.set_team_co_commissioner(
        league["id"], member["id"], enabled=True, actor_sub="life-comm"
    )
    start_league_delete(league["id"], actor_sub="life-comm", confirm_name="Withdraw Delete")
    cancelled = cancel_league_delete(league["id"], actor_sub="life-co-c")
    assert cancelled["pending"] is False
    assert storage.get_league(league["id"]) is not None
    again = start_league_delete(league["id"], actor_sub="life-comm", confirm_name="Withdraw Delete")
    assert again["deleted"] is False


def test_home_surfaces_pending_delete(hub_db):
    league, team = _seed_league("Home Delete")
    member = storage.join_league("life-co-d", league["room_code"], "Co D")
    storage.set_team_co_commissioner(
        league["id"], member["id"], enabled=True, actor_sub="life-comm"
    )
    start_league_delete(league["id"], actor_sub="life-comm", confirm_name="Home Delete")
    ctx = {
        "mode": "league",
        "league_id": league["id"],
        "league_name": league["name"],
        "team_id": member["id"],
        "team_name": member["name"],
        "is_commissioner": True,
        "draft_completed": False,
        "league_status": "setup",
        "rules": league["rules"],
        "season": 2026,
        "workspace_id": storage.roster_workspace_for_league(league),
    }
    payload = build_league_home(ctx, include_week=False)
    ids = [item["id"] for item in payload["actions"]]
    assert "delete_league" in ids
    action = next(item for item in payload["actions"] if item["id"] == "delete_league")
    assert action["href"] == "office-access"
    assert action["severity"] == "high"

    starter_ctx = {**ctx, "team_id": team["id"], "team_name": team["name"]}
    starter_home = build_league_home(starter_ctx, include_week=False)
    starter_ids = [item["id"] for item in starter_home["actions"]]
    assert "delete_league_wait" in starter_ids


def test_delete_clears_contract_history(hub_db):
    league, _team = _seed_league("History Delete")
    storage.replace_league_contract_season_source(
        league["id"],
        2025,
        [
            {
                "owner_label": "Comm",
                "player_name": "Gone Player",
                "position": "RB",
                "base_salary": 12,
                "cap_hit": 12,
            }
        ],
        source_kind="import",
    )
    lid = league["id"]
    start_league_delete(lid, actor_sub="life-comm", confirm_name="History Delete")
    assert storage.get_league(lid) is None
    assert storage.list_league_contract_rows(lid) == []
